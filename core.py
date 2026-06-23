# -*- coding: utf-8 -*-
"""
core.py — движок генерации документов из docx-шаблонов.
Зависимости: jinja2, openpyxl.

Разметка в шаблонах (.docx):
  {{ переменная }}                — подстановка значения (перенос строки \n работает)
  {%p if условие %} ... {%p endif %}   — условный блок: тег занимает ОТДЕЛЬНЫЙ абзац,
                                          абзац-маркер удаляется из результата
  {%tr for у in услуги %} ... {%tr endfor %} — цикл по строкам таблицы:
                                          теги в отдельных строках до/после образца
"""
import io, os, re, json, zipfile, datetime
from decimal import Decimal, ROUND_HALF_UP
from xml.sax.saxutils import escape as _xml_escape

import jinja2

# ----------------------------------------------------------------------------
# 1. ЧИСЛА И ДАТЫ ПРОПИСЬЮ
# ----------------------------------------------------------------------------
_UNITS_M = ["", "один", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
_UNITS_F = ["", "одна", "две", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
_TEENS   = ["десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать",
            "пятнадцать", "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать"]
_TENS    = ["", "", "двадцать", "тридцать", "сорок", "пятьдесят",
            "шестьдесят", "семьдесят", "восемьдесят", "девяносто"]
_HUNDREDS = ["", "сто", "двести", "триста", "четыреста", "пятьсот",
             "шестьсот", "семьсот", "восемьсот", "девятьсот"]


def _plural(n, one, few, many):
    n = abs(int(n)) % 100
    if 11 <= n <= 19:
        return many
    n %= 10
    if n == 1:
        return one
    if 2 <= n <= 4:
        return few
    return many


def _triple(n, feminine=False):
    units = _UNITS_F if feminine else _UNITS_M
    words = []
    h, rest = divmod(n, 100)
    if h:
        words.append(_HUNDREDS[h])
    if 10 <= rest <= 19:
        words.append(_TEENS[rest - 10])
    else:
        t, u = divmod(rest, 10)
        if t:
            words.append(_TENS[t])
        if u:
            words.append(units[u])
    return " ".join(words)


def chislo_propisyu(n, feminine=False):
    n = int(n)
    if n == 0:
        return "ноль"
    scales = [(10 ** 9, ("миллиард", "миллиарда", "миллиардов"), False),
              (10 ** 6, ("миллион", "миллиона", "миллионов"), False),
              (10 ** 3, ("тысяча", "тысячи", "тысяч"), True)]
    words = []
    for div, forms, fem in scales:
        q, n = divmod(n, div)
        if q:
            words.append(_triple(q, feminine=fem))
            words.append(_plural(q, *forms))
    if n:
        words.append(_triple(n, feminine=feminine))
    return " ".join(w for w in words if w)


def rub_propisyu(amount):
    amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    rub = int(amount)
    kop = int((amount - rub) * 100)
    s = chislo_propisyu(rub)
    s = s[:1].upper() + s[1:]
    return f"{s} {_plural(rub, 'рубль', 'рубля', 'рублей')} {kop:02d} {_plural(kop, 'копейка', 'копейки', 'копеек')}"


def money_fmt(amount):
    amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{amount:,.2f}".replace(",", " ").replace(".", ",")


_UNITS_G = ["", "одного", "двух", "трёх", "четырёх", "пяти", "шести", "семи", "восьми", "девяти"]
_TEENS_G = ["десяти", "одиннадцати", "двенадцати", "тринадцати", "четырнадцати",
            "пятнадцати", "шестнадцати", "семнадцати", "восемнадцати", "девятнадцати"]
_TENS_G  = ["", "", "двадцати", "тридцати", "сорока", "пятидесяти",
            "шестидесяти", "семидесяти", "восьмидесяти", "девяноста"]


def dnej_propisyu(n):
    n = int(n)
    if n == 100:
        return "ста"
    if 10 <= n <= 19:
        return _TEENS_G[n - 10]
    t, u = divmod(n, 10)
    parts = []
    if t:
        parts.append(_TENS_G[t])
    if u:
        parts.append(_UNITS_G[u])
    return " ".join(parts) if parts else "ноля"


_MONTHS_G = ["", "января", "февраля", "марта", "апреля", "мая", "июня",
             "июля", "августа", "сентября", "октября", "ноября", "декабря"]


def date_ru(d):
    if isinstance(d, str) or d is None:
        return d or ""
    return f"«{d.day:02d}» {_MONTHS_G[d.month]} {d.year} г."


# ----------------------------------------------------------------------------
# 2. ПРОВЕРКА РЕКВИЗИТОВ (контрольные цифры)
# ----------------------------------------------------------------------------
def _digits(s):
    return re.sub(r"\D", "", str(s or ""))


def check_inn(inn):
    """None, если ИНН корректен; иначе текст проблемы."""
    inn = _digits(inn)
    if not inn:
        return None
    if len(inn) == 10:
        w = [2, 4, 10, 3, 5, 9, 4, 6, 8]
        if sum(int(a) * b for a, b in zip(inn, w)) % 11 % 10 == int(inn[9]):
            return None
        return "не сходится контрольная цифра ИНН"
    if len(inn) == 12:
        w1 = [7, 2, 4, 10, 3, 5, 9, 4, 6, 8]
        w2 = [3, 7, 2, 4, 10, 3, 5, 9, 4, 6, 8]
        ok = (sum(int(a) * b for a, b in zip(inn, w1)) % 11 % 10 == int(inn[10]) and
              sum(int(a) * b for a, b in zip(inn, w2)) % 11 % 10 == int(inn[11]))
        return None if ok else "не сходится контрольная цифра ИНН"
    return "ИНН должен содержать 10 (организация) или 12 (физлицо/ИП) цифр"


def check_ogrn(ogrn):
    ogrn = _digits(ogrn)
    if not ogrn:
        return None
    if len(ogrn) == 13:
        return None if int(ogrn[:12]) % 11 % 10 == int(ogrn[12]) \
            else "не сходится контрольная цифра ОГРН"
    if len(ogrn) == 15:
        return None if int(ogrn[:14]) % 13 % 10 == int(ogrn[14]) \
            else "не сходится контрольная цифра ОГРНИП"
    return "ОГРН — 13 цифр, ОГРНИП — 15 цифр"


def check_bik(bik):
    bik = _digits(bik)
    if not bik:
        return None
    if len(bik) != 9 or not bik.startswith("04"):
        return "БИК — 9 цифр, начинается с 04"
    return None


def validate_party(p, who):
    """Список предупреждений по реквизитам стороны."""
    out = []
    for field, fn in [("инн", check_inn), ("огрн", check_ogrn), ("бик", check_bik)]:
        msg = fn(p.get(field))
        if msg:
            out.append(f"{who}: {msg} ({p.get(field)})")
    return out


# ----------------------------------------------------------------------------
# 3. РЕНДЕР DOCX-ШАБЛОНОВ
# ----------------------------------------------------------------------------
_RX_TAG = re.compile(r"(\{[\{%])(.*?)([\}%]\})", re.DOTALL)
_RX_RUNBREAK = re.compile(r"</w:t>.*?<w:t(?: [^>]*)?>", re.DOTALL)
_RX_P_TAG = re.compile(
    r"<w:p\b(?:(?!</w:p>).)*?\{%p((?:(?!%\}).)*?)%\}(?:(?!</w:p>).)*?</w:p>", re.DOTALL)
_RX_TR_TAG = re.compile(
    r"<w:tr\b(?:(?!</w:tr>).)*?\{%tr((?:(?!%\}).)*?)%\}(?:(?!</w:tr>).)*?</w:tr>", re.DOTALL)


def _merge_split_tags(xml):
    def fix(m):
        inner = _RX_RUNBREAK.sub("", m.group(2))
        return m.group(1) + inner + m.group(3)
    return _RX_TAG.sub(fix, xml)


def _prepare_xml(xml):
    xml = _merge_split_tags(xml)
    xml = _RX_TR_TAG.sub(lambda m: "{%" + m.group(1) + "%}", xml)
    xml = _RX_P_TAG.sub(lambda m: "{%" + m.group(1) + "%}", xml)
    return xml


def _finalize(value):
    s = _xml_escape(str(value))
    return s.replace("\n", '</w:t><w:br/><w:t xml:space="preserve">')


_env = jinja2.Environment(undefined=jinja2.StrictUndefined, finalize=_finalize,
                          autoescape=False)


def render_docx_bytes(template_bytes, context):
    """Заполняет docx-шаблон (bytes) и возвращает результат (bytes)."""
    with zipfile.ZipFile(io.BytesIO(template_bytes)) as zin:
        items = {n: zin.read(n) for n in zin.namelist()}
    targets = [n for n in items if re.fullmatch(
        r"word/(document|header\d*|footer\d*)\.xml", n)]
    for name in targets:
        xml = _prepare_xml(items[name].decode("utf-8"))
        try:
            xml = _env.from_string(xml).render(**context)
        except jinja2.TemplateError as e:
            raise RuntimeError(f"ошибка шаблона: {e}") from e
        items[name] = xml.encode("utf-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in items.items():
            zout.writestr(name, data)
    return buf.getvalue()


def render_docx(template_path, context, out_path):
    with open(template_path, "rb") as f:
        data = render_docx_bytes(f.read(), context)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with open(out_path, "wb") as f:
        f.write(data)
    return out_path


# ----------------------------------------------------------------------------
# 4. КОНТЕКСТ (поля + вычисляемые значения)
# ----------------------------------------------------------------------------

def num_str(d):
    """Decimal -> строка без хвостовых нулей: 20 -> «20», 2.50 -> «2,5»."""
    s = f"{Decimal(str(d)):f}"
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return (s or "0").replace(".", ",")

def rekvizity_text(p):
    lines = [p.get("наименование", "")]
    if p.get("инн"):
        s = "ИНН " + p["инн"]
        if p.get("кпп"):
            s += ", КПП " + p["кпп"]
        lines.append(s)
    if p.get("огрн"):
        lines.append(("ОГРНИП " if p.get("тип") == "ИП" else "ОГРН ") + p["огрн"])
    if p.get("паспорт"):
        lines.append("Паспорт: " + p["паспорт"])
    if p.get("адрес"):
        lines.append("Адрес: " + p["адрес"])
    if p.get("счет"):
        lines.append("Р/с " + p["счет"])
    if p.get("банк"):
        lines.append("Банк: " + p["банк"])
    if p.get("бик"):
        s = "БИК " + p["бик"]
        if p.get("корсчет"):
            s += ", к/с " + p["корсчет"]
        lines.append(s)
    if p.get("телефон"):
        lines.append("Тел.: " + p["телефон"])
    if p.get("email"):
        lines.append("E-mail: " + p["email"])
    return "\n".join(l for l in lines if l)


def short_fio(fio):
    parts = str(fio or "").split()
    if len(parts) >= 2 and "«" not in str(fio):
        return parts[0] + " " + ".".join(w[0] for w in parts[1:]) + "."
    return str(fio or "")


def build_context(d, active_doc=None):
    """active_doc: 'прил'|'счет'|'акт'|'дог' — какой документ сейчас рендерится;
    его номер/дата попадают в договор_номер/договор_дата. Если не задан или
    у документа нет своего номера — берётся общий d['договор_номер'] (совместимость)."""
    зак, исп = d["заказчик"], d["исполнитель"]
    исп_тип = исп.get("тип") or "Самозанятый"
    самозанятый = исп_тип == "Самозанятый"

    # номера/даты документов: каждый свой, с откатом на общий «договорный»
    общий_номер = d.get("договор_номер") or d.get("дог_номер") or ""
    общая_дата = d.get("договор_дата") or d.get("дог_дата")

    def ном(тип):
        return d.get(f"{тип}_номер") or общий_номер

    def дат(тип):
        return d.get(f"{тип}_дата") or общая_дата

    прил_номер, прил_дата = ном("прил") or "1", дат("прил")
    счет_номер, счет_дата = ном("счет"), дат("счет")
    акт_номер, акт_дата = ном("акт"), дат("акт")
    # номер/дата самого договора (тип «дог») — для строки «к Договору №…»
    дог_ном_отд = d.get("дог_номер") or ""
    дог_дата_отд = d.get("дог_дата")

    # «активный» документ задаёт договор_номер/дата (для имени файла и общих тегов)
    akt_map = {"прил": (ном("прил"), дат("прил")), "счет": (счет_номер, счет_дата),
               "акт": (акт_номер, акт_дата), "дог": (общий_номер, общая_дата)}
    дн_номер, дн_дата = akt_map.get(active_doc, (общий_номер, общая_дата))
    дн_номер = дн_номер or общий_номер

    услуги_ctx, итого = [], Decimal("0")
    for i, u in enumerate(u for u in d.get("услуги", [])
                          if str(u.get("наименование", "")).strip()):
        колво = Decimal(str(u.get("колво") or 1))
        цена = u.get("цена")
        if цена in (None, ""):
            цена = u.get("стоимость") or 0
        цена = Decimal(str(цена))
        сумма = (цена * колво).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        итого += сумма
        услуги_ctx.append({
            "номер": i + 1,
            "наименование": str(u.get("наименование", "")),
            "колво": num_str(колво),
            "цена": money_fmt(цена),
            "стоимость": money_fmt(сумма),
            "начало": date_ru(u.get("начало") or d.get("услуги_начало")),
            "окончание": date_ru(u.get("окончание") or d.get("услуги_окончание")),
            "требования": str(u.get("требования") or "—"),
        })
    стоимость = Decimal(str(d.get("стоимость") or 0)) or итого

    ндс_ставка = Decimal(str(d.get("ндс_ставка") or 0))
    ндс_сумма = (стоимость * ндс_ставка / 100).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP)
    итого_с_ндс = стоимость + ндс_сумма

    ctx = {
        "договор_номер": дн_номер,
        "договор_дата": date_ru(дн_дата),
        "город": d.get("город", ""),
        "услуги_начало": date_ru(d.get("услуги_начало")),
        "услуги_окончание": date_ru(d.get("услуги_окончание")),

        "заказчик_наименование": зак.get("наименование", ""),
        "заказчик_должность_род": зак.get("должность_род") or "Генерального директора",
        "заказчик_подписант_род": зак.get("подписант_род", ""),
        "заказчик_основание_род": зак.get("основание_род") or "Устава",
        "заказчик_подписант_кратко": short_fio(зак.get("подписант") or зак.get("подписант_род")),
        "заказчик_инн": зак.get("инн", ""), "заказчик_кпп": зак.get("кпп", ""),
        "заказчик_огрн": зак.get("огрн", ""),
        "заказчик_адрес": зак.get("адрес", ""),
        "заказчик_реквизиты": rekvizity_text(зак),

        "исполнитель_фио": исп.get("наименование", ""),
        "исполнитель_фио_кратко": short_fio(исп.get("подписант") or исп.get("наименование")),
        "исполнитель_подписант_кратко": short_fio(исп.get("подписант") or исп.get("наименование")),
        "исполнитель_тип": исп_тип,
        "исполнитель_инн": исп.get("инн", ""), "исполнитель_кпп": исп.get("кпп", ""),
        "исполнитель_огрн": исп.get("огрн", ""),
        "исполнитель_должность_род": исп.get("должность_род") or "Генерального директора",
        "исполнитель_подписант_род": исп.get("подписант_род", ""),
        "исполнитель_основание_род": исп.get("основание_род") or "Устава",
        "исполнитель_паспорт": исп.get("паспорт", ""),
        "исполнитель_адрес": исп.get("адрес", ""),
        "исполнитель_банк": исп.get("банк", ""), "исполнитель_бик": исп.get("бик", ""),
        "исполнитель_счет": исп.get("счет", ""),
        "исполнитель_корсчет": исп.get("корсчет", ""),
        "исполнитель_реквизиты": rekvizity_text(исп),

        "стоимость_цифрами": money_fmt(стоимость),
        "стоимость_прописью": rub_propisyu(стоимость),
        "ндс_ставка": num_str(ндс_ставка),
        "ндс_сумма": money_fmt(ндс_сумма),
        "итого_с_ндс_цифрами": money_fmt(итого_с_ндс),
        "итого_с_ндс_прописью": rub_propisyu(итого_с_ндс),

        "услуги": услуги_ctx,
        "блок_ис": bool(d.get("блок_ис", True)),

        "акт_номер": акт_номер or дн_номер,
        "акт_дата": date_ru(акт_дата or d.get("услуги_окончание")),
        "счет_номер": счет_номер or дн_номер,
        "счет_дата": date_ru(счет_дата or общая_дата),
        "прил_номер": прил_номер,
        "прил_дата": date_ru(прил_дата or общая_дата),
        "договор_номер_дог": (дог_ном_отд or "______"),
        "договор_дата_дог": (date_ru(дог_дата_отд) if дог_дата_отд else "______"),

        "оферта_срок_оплаты": d.get("оферта_срок_оплаты")
            or "в течение 5 (пяти) рабочих дней с даты выставления Счета",
        "оферта_срок_работ": d.get("оферта_срок_работ")
            or "в течение 10 (десяти) рабочих дней с даты внесения аванса",
        "оферта_результат": d.get("оферта_результат")
            or "результат работ, указанных в Счете",
        "оферта_формат": d.get("оферта_формат") or "ссылкой на облачное хранилище",
        "оферта_срок_размещения": d.get("оферта_срок_размещения") or "______",
        "оферта_отношения_с": (date_ru(d["оферта_отношения_с"])
            if isinstance(d.get("оферта_отношения_с"),
                          (datetime.date, datetime.datetime))
            else (d.get("оферта_отношения_с") or "______")),
        "оферта_место": d.get("оферта_место")
            or "г. Екатеринбург, ул. Сакко и Ванцетти, д. 61",
        "аванс_процент": num_str(d.get("аванс_процент") or 50),
        "исполнитель_email_или_прочерк": (исп.get("email") or "________"),
        "оферта_основание_номер": d.get("оферта_основание_номер") or "______",
        "оферта_основание_дата": (date_ru(d.get("оферта_основание_дата"))
                                  if d.get("оферта_основание_дата") else "______"),
        "прил_оплата": d.get("прил_оплата") or (
            "Заказчик обязуется оплатить Услуги в размере 100% их стоимости до "
            "начала оказания Услуг путем перечисления денежных средств на банковский "
            "счет Исполнителя на основании счета на оплату, выставленного "
            "Исполнителем. Акт оказанных услуг направляется Исполнителем Заказчику "
            "по факту оказания Услуг."),
        "ндс_строка": d.get("ндс_строка")
            or ("без НДС (Подрядчик применяет НПД)" if самозанятый
                else f"НДС {num_str(ндс_ставка)}% — {money_fmt(ндс_сумма)} руб."
                if ндс_ставка else "без НДС"),
        "ндс_предложение": d.get("ндс_предложение")
            or ("НДС не облагается, поскольку Подрядчик является плательщиком налога "
                "на профессиональный доход." if самозанятый else
                f"В том числе НДС {num_str(ндс_ставка)}% — {money_fmt(ндс_сумма)} руб."
                if ндс_ставка else "НДС не облагается."),
    }
    # совместимость со старым комплектом (_архив): дни/штрафы/блоки
    for k, v in {"справка_дней": 3, "пояснения_дней": 3, "акт_дней": 5,
                 "приемка_дней": 5, "устранение_дней": 5, "предоплата_дней": 5,
                 "постоплата_дней": 5, "расторжение_дней": 10, "претензия_дней": 10,
                 "фм_уведомление_дней": 10}.items():
        n = int(d.get(k, v))
        ctx[k], ctx[k + "_проп"] = n, dnej_propisyu(n)
    ctx.update({"предоплата_процент": d.get("предоплата_процент", 100),
                "постоплата_процент": d.get("постоплата_процент", 0),
                "штраф_чек_процент": d.get("штраф_чек_процент", 20),
                "штраф_уведомление": money_fmt(d.get("штраф_уведомление", 10000)),
                "штраф_уведомление_проп": rub_propisyu(d.get("штраф_уведомление", 10000)),
                "блок_нда": bool(d.get("блок_нда")), "блок_фм": bool(d.get("блок_фм")),
                "блок_согласие": bool(d.get("блок_согласие")),
                "нда_лет": 3, "нда_лет_проп": "три",
                "фм_месяцев": 2, "фм_месяцев_проп": "два",
                "согласие_до": date_ru(d.get("согласие_до") or d.get("услуги_окончание"))})
    n = 6
    if ctx["блок_нда"]:
        n += 1; ctx["с_нда"] = n
    if ctx["блок_фм"]:
        n += 1; ctx["с_фм"] = n
    ctx.update({"с_раст": n + 1, "с_спор": n + 2, "с_проч": n + 3,
                "с_прил": n + 4, "с_рекв": n + 5})
    return ctx


# ----------------------------------------------------------------------------
# 5. КОМПЛЕКТЫ И ГЕНЕРАЦИЯ
# ----------------------------------------------------------------------------
TPL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")
MANIFEST = "манифест.json"
SINGLE_SET = "Самозанятые"  # комплект по умолчанию для варианта без выбора


def local_sets():
    """Комплекты из локальной папки templates/ (резерв без Диска)."""
    out = {}
    if not os.path.isdir(TPL_DIR):
        return out
    for name in sorted(os.listdir(TPL_DIR)):
        p = os.path.join(TPL_DIR, name)
        if name.startswith("_") or not os.path.isdir(p):
            continue
        out[name] = sorted(f for f in os.listdir(p) if f.endswith(".docx"))
    return out


def read_manifest(raw, set_name):
    """Манифест комплекта; при отсутствии — разумные значения по имени папки.
    Дополнительно:
      - "документы": {"имя файла.docx": "счет"|"акт"|"прил"|"дог"|"нда"} —
        тип документа берётся отсюда, а не угадывается по имени;
      - "выбор_типа": true — в комплекте есть переключатель ИП↔ЮЛ;
      - "вид": "нда" — особый комплект НДА (минимум полей)."""
    m = {}
    if raw:
        try:
            m = json.loads(raw.decode("utf-8") if isinstance(raw, bytes) else raw)
        except Exception:
            m = {}
    low = set_name.strip().lower()
    тип = m.get("тип_исполнителя") or (
        "ИП" if low == "ип" else
        "ЮЛ" if "юл" in low else
        "Самозанятый")
    префикс = m.get("префикс") or {"ИП": "ИП", "ЮЛ": "ЮЛ"}.get(тип, "СЗ")
    return {"тип_исполнителя": тип, "префикс": префикс,
            "ндс": bool(m.get("ндс", тип == "ЮЛ")),
            "документы": m.get("документы", {}),
            "выбор_типа": bool(m.get("выбор_типа", False)),
            "вид": m.get("вид", "обычный")}


def safe_name(s):
    for ch in '/\\:*?"<>|':
        s = s.replace(ch, "-")
    return s.strip()


def scrub_pii(text):
    """Удаляет потенциальные персональные данные из текста (для логов/сообщений
    об ошибках): длинные числа (счета, ИНН, ОГРН), e-mail, паспортные серии.
    Не идеальная защита, а снижение риска утечки ПД в логи Render/Streamlit."""
    s = str(text)
    s = re.sub(r"[\w.+-]+@[\w-]+\.[\w.-]+", "[email]", s)          # e-mail
    s = re.sub(r"\b\d{9,20}\b", "[номер]", s)                     # счета/ИНН/ОГРН
    s = re.sub(r"\b\d{2}\s?\d{2}\s?№?\s?\d{6}\b", "[паспорт]", s)  # паспорт
    return s


def _doc_type_of(fname):
    low = os.path.splitext(fname)[0].lower()
    if "приложен" in low:
        return "прил"
    if "счёт" in low or "счет" in low or "оферт" in low:
        return "счет"
    if "акт" in low:
        return "акт"
    return "дог"


def generate_files(d, out_dir, templates, doc_types=None):
    """templates: список (имя_файла, bytes). Каждый файл рендерится со своим
    номером/датой (по типу документа). doc_types: {имя_файла: тип} из манифеста;
    если для файла типа нет — определяется по имени. Возвращает пути результатов."""
    doc_types = doc_types or {}
    paths = []
    os.makedirs(out_dir, exist_ok=True)
    for fname, data in templates:
        тип = doc_types.get(fname) or _doc_type_of(fname)
        ctx = build_context(d, active_doc=тип)
        stem = os.path.splitext(fname)[0]
        ном = ctx["договор_номер"]
        suffix = f" №{ном}" if ном else ""
        out = os.path.join(out_dir, safe_name(stem + suffix) + ".docx")
        try:
            result = render_docx_bytes(data, ctx)
        except RuntimeError as e:
            raise RuntimeError(f"«{fname}»: {e}") from e
        with open(out, "wb") as f:
            f.write(result)
        paths.append(out)
    return paths


def test_data(тип="Самозанятый"):
    """Тестовые данные для кнопки «Проверить комплект»."""
    return {
        "договор_номер": "ТЕСТ-1",
        "договор_дата": datetime.date.today(),
        "город": "Екатеринбург",
        "услуги_начало": datetime.date.today(),
        "услуги_окончание": datetime.date.today(),
        "заказчик": {"тип": "АО", "наименование": "АО «Тест»", "инн": "9705120864",
                     "кпп": "770401001", "огрн": "1187746637143",
                     "адрес": "г. Москва", "подписант": "Тестов Тест Тестович",
                     "подписант_род": "Тестова Теста Тестовича"},
        "исполнитель": {"тип": тип, "наименование": "Пример Примеров Примерович",
                        "инн": "772212345678", "огрн": "326665800012345",
                        "кпп": "667101001", "адрес": "г. Екатеринбург",
                        "счет": "40802810600000098765", "банк": "ООО «Банк Точка»",
                        "бик": "044525104", "корсчет": "30101810745374525104",
                        "паспорт": "00 00 № 000000",
                        "подписант": "Пример Примеров Примерович",
                        "подписант_род": "Примерова Примера Примеровича"},
        "услуги": [{"наименование": "Тестовая работа", "колво": 1, "цена": 1000}],
        "ндс_ставка": 20,
    }


# ----------------------------------------------------------------------------
# 6. БАЗА КОНТРАГЕНТОВ (Excel): записи, архив версий, журнал
# ----------------------------------------------------------------------------
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "контрагенты.xlsx")
PARTY_COLS = ["тип", "наименование", "инн", "кпп", "огрн", "адрес",
              "должность_род", "подписант_род", "подписант", "основание_род",
              "паспорт", "счет", "банк", "бик", "корсчет", "телефон", "email"]
ARCHIVE_COLS = PARTY_COLS + ["заменено", "кем"]
JOURNAL_COLS = ["создано", "договор_номер", "договор_дата", "комплект", "заказчик",
                "исполнитель", "сумма", "документы", "кто"]


def _norm(v):
    return str(v or "").strip()


def party_key(p):
    return _digits(p.get("инн"))


_STATUS = {"самозанятый", "ип", "юл"}  # значимые для раздельных записей статусы
_STATUS_LABEL = {"Самозанятый", "ИП", "ЮЛ"}  # как выбирается тип комплекта


def _same_party(a, b):
    """Та же запись: совпадает ИНН. Тип-статус (самозанятый/ИП/ЮЛ) различает
    записи только если у обеих сторон он задан из этого набора и различается —
    это сценарий смены статуса. Организационные формы (АО, ООО и пр.) и пустой
    тип на совпадение по ИНН не влияют."""
    if party_key(a) != party_key(b) or not party_key(a):
        return False
    ta, tb = _norm(a.get("тип")).lower(), _norm(b.get("тип")).lower()
    if ta in _STATUS and tb in _STATUS and ta != tb:
        return False
    return True


def _open_db():
    import openpyxl
    if os.path.exists(DB_PATH):
        wb = openpyxl.load_workbook(DB_PATH)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "Контрагенты"
        wb.active.append(PARTY_COLS)
    if "Контрагенты" not in wb.sheetnames:
        wb.create_sheet("Контрагенты").append(PARTY_COLS)
    if "Архив" not in wb.sheetnames:
        wb.create_sheet("Архив").append(ARCHIVE_COLS)
    if "Журнал" not in wb.sheetnames:
        wb.create_sheet("Журнал").append(JOURNAL_COLS)
    ws = wb["Журнал"]  # миграция старого журнала на новые колонки
    if ws.max_column < len(JOURNAL_COLS):
        for i, name in enumerate(JOURNAL_COLS, 1):
            ws.cell(row=1, column=i, value=name)
    return wb


def _row_to_party(row):
    return dict(zip(PARTY_COLS, [_norm(v) for v in row[:len(PARTY_COLS)]]))


# Заказчики Точки — заносятся в базу при первом запуске (если их там нет)
DEFAULT_CUSTOMERS = [
    {"тип": "ЮЛ", "наименование": "Акционерное общество «Точка»",
     "инн": "9705120864", "кпп": "770401001", "огрн": "1187746637143",
     "адрес": ("119002, г. Москва, вн.тер.г. Муниципальный округ Хамовники, "
               "пер. Староконюшенный, д. 10; адрес для корреспонденции: 620014, "
               "г. Екатеринбург, ул. Сакко и Ванцетти, д. 61"),
     "подписант": "Гаврелюк Алёна Борисовна",
     "подписант_род": "Гаврелюк Алёны Борисовны",
     "основание_род": "доверенности № 36 от 06.03.2025",
     "счет": "40702810820000000001", "банк": "ООО «Банк Точка»",
     "бик": "044525104", "корсчет": "30101810745374525104"},
    {"тип": "ЮЛ",
     "наименование": "Общество с ограниченной ответственностью «Банк Точка»",
     "инн": "9721194461", "кпп": "772301001/997950001", "огрн": "1237700005157",
     "адрес": ("109044, г. Москва, вн.тер.г. муниципальный округ Южнопортовый, "
               "пер. 3-й Крутицкий, д. 11, помещ. 7Н; адрес для корреспонденции: "
               "620014, г. Екатеринбург, ул. Сакко и Ванцетти, д. 61"),
     "подписант": "Лучшева Любовь Александровна",
     "подписант_род": "Лучшевой Любови Александровны",
     "основание_род": "Доверенности № 44 от 25.04.2025",
     "банк": "ОКЦ № 1 ГУ Банка России по ЦФО", "бик": "044525104",
     "корсчет": "30101810745374525104"},
]


def seed_customers(user="система"):
    """Заносит заказчиков Точки в базу, если их там ещё нет (по ИНН)."""
    added = 0
    for c in DEFAULT_CUSTOMERS:
        if not find_party({"инн": c["инн"]}):
            save_party(dict(c), user=user)
            added += 1
    return added


def load_parties():
    if not os.path.exists(DB_PATH):
        return []
    ws = _open_db()["Контрагенты"]
    return [_row_to_party(r) for r in ws.iter_rows(min_row=2, values_only=True)
            if any(r)]


def find_party(p):
    """Текущая запись с тем же ИНН (или None). См. _same_party про смену статуса."""
    if not party_key(p):
        return None
    for q in load_parties():
        if _same_party(q, p):
            return q
    return None


def party_diff(old, new):
    """[(поле, было, стало)] по значимым отличиям. Поле «тип» не сравниваем:
    смена налогового статуса уже разводится в отдельную запись (_same_party),
    а смена ярлыка формы (АО↔ЮЛ из разных мест ввода) изменением не считается."""
    out = []
    for k in PARTY_COLS:
        if k == "тип":
            continue
        a, b = _norm(old.get(k)), _norm(new.get(k))
        if a != b and (a or b):
            out.append((k, a or "—", b or "—"))
    return out


def save_party(p, user="", overwrite=True):
    """Сохраняет контрагента. Если запись с тем же ИНН+типом изменилась и
    overwrite=True — старая версия уходит в «Архив». Возврат: 'new'|'updated'|
    'same'|'skipped'."""
    if not _digits(p.get("инн")):
        return "skipped"
    wb = _open_db()
    ws, wa = wb["Контрагенты"], wb["Архив"]
    for row in ws.iter_rows(min_row=2):
        cur = _row_to_party([c.value for c in row])
        if _same_party(cur, p):
            if not party_diff(cur, p):
                return "same"
            if not overwrite:
                return "skipped"
            stamp = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
            wa.append([cur.get(c, "") for c in PARTY_COLS] + [stamp, user])
            # тип не перезаписываем «вслепую»: если в форме пусто или общий ярлык
            # (ЮЛ), а в базе был осмысленный (АО/ООО) — оставляем прежний
            new_type = _norm(p.get("тип"))
            if not new_type or (new_type.lower() == "юл" and _norm(cur.get("тип"))):
                p = dict(p, тип=cur.get("тип"))
            for i, c in enumerate(PARTY_COLS):
                row[i].value = _norm(p.get(c))
            wb.save(DB_PATH)
            return "updated"
    ws.append([_norm(p.get(c)) for c in PARTY_COLS])
    wb.save(DB_PATH)
    return "new"


def party_history(p):
    """Версии контрагента из «Архива», новые сверху."""
    out = []
    if not os.path.exists(DB_PATH):
        return out
    ws = _open_db()["Архив"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        rec = dict(zip(ARCHIVE_COLS, [_norm(v) for v in r]))
        if _same_party(rec, p):
            out.append(rec)
    return list(reversed(out))


def restore_party(version, user=""):
    """Возвращает архивную версию в «Контрагенты»; текущая уходит в архив."""
    p = {k: version.get(k, "") for k in PARTY_COLS}
    return save_party(p, user=user, overwrite=True)


def _primary_number(d):
    """Главный номер для журнала/имени: приложение → счёт → акт → общий."""
    for тип in ("прил", "счет", "акт", "дог"):
        v = _norm(d.get(f"{тип}_номер"))
        if v:
            return v
    return _norm(d.get("договор_номер"))


def _primary_date(d):
    for тип in ("прил", "счет", "акт", "дог"):
        if d.get(f"{тип}_номер") and d.get(f"{тип}_дата"):
            return d.get(f"{тип}_дата")
    return d.get("договор_дата") or d.get("прил_дата") or d.get("счет_дата") \
        or d.get("акт_дата")


def append_journal(d, файлы, комплект="", user=""):
    wb = _open_db()
    ws = wb["Журнал"]
    сумма = d.get("стоимость") or sum(
        float(u.get("цена") or u.get("стоимость") or 0) * float(u.get("колво") or 1)
        for u in d.get("услуги", []))
    ws.append([datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
               _primary_number(d), date_ru(_primary_date(d)), комплект,
               d["заказчик"].get("наименование", ""),
               d["исполнитель"].get("наименование", ""),
               money_fmt(сумма), ", ".join(файлы), user])
    wb.save(DB_PATH)


def journal_numbers():
    if not os.path.exists(DB_PATH):
        return []
    ws = _open_db()["Журнал"]
    return [_norm(r[1]) for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]


def next_contract_number(prefix):
    """Следующий номер по комплекту: СЗ-1, СЗ-2, …"""
    rx = re.compile(rf"^{re.escape(prefix)}-(\d+)$", re.IGNORECASE)
    nums = [int(m.group(1)) for n in journal_numbers() if (m := rx.match(n))]
    return f"{prefix}-{max(nums) + 1 if nums else 1}"
