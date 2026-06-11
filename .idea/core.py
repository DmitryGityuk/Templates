# -*- coding: utf-8 -*-
"""
core.py — движок генерации документов из docx-шаблонов.
Зависимости: jinja2, openpyxl (стандартные, ставятся через pip).

Разметка в шаблонах (.docx):
  {{ переменная }}                — подстановка значения (перенос строки \n работает)
  {%p if условие %} ... {%p endif %}   — условный блок: тег занимает ОТДЕЛЬНЫЙ абзац,
                                          абзац-маркер удаляется из результата
  {%tr for у in услуги %} ... {%tr endfor %} — цикл по строкам таблицы:
                                          тег ставится в ячейку строки-шаблона
"""
import io, os, re, zipfile, datetime
from decimal import Decimal, ROUND_HALF_UP
from xml.sax.saxutils import escape as _xml_escape

import jinja2

# ----------------------------------------------------------------------------
# 1. ЧИСЛА И ДАТЫ ПРОПИСЬЮ
# ----------------------------------------------------------------------------
_UNITS_M = ["", "один", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
_UNITS_F = ["", "одна", "две", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
_TEENS   = ["десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать",
            "пятнадцать", "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать"]
_TENS    = ["", "", "двадцать", "тридцать", "сорок", "пятьдесят",
            "шестьдесят", "семьдесят", "восемьдесят", "девяносто"]
_HUNDREDS = ["", "сто", "двести", "триста", "четыреста", "пятьсот",
             "шестьсот", "семьсот", "восемьсот", "девятьсот"]


def _plural(n, one, few, many):
    n = abs(int(n)) % 100
    if 11 <= n <= 19:
        return many
    n %= 10
    if n == 1:
        return one
    if 2 <= n <= 4:
        return few
    return many


def _triple(n, feminine=False):
    units = _UNITS_F if feminine else _UNITS_M
    words = []
    h, rest = divmod(n, 100)
    if h:
        words.append(_HUNDREDS[h])
    if 10 <= rest <= 19:
        words.append(_TEENS[rest - 10])
    else:
        t, u = divmod(rest, 10)
        if t:
            words.append(_TENS[t])
        if u:
            words.append(units[u])
    return " ".join(words)


def chislo_propisyu(n, feminine=False):
    """Целое число прописью (именительный падеж), 0..999 млрд."""
    n = int(n)
    if n == 0:
        return "ноль"
    scales = [  # (делитель, формы, женский род?)
        (10 ** 9, ("миллиард", "миллиарда", "миллиардов"), False),
        (10 ** 6, ("миллион", "миллиона", "миллионов"), False),
        (10 ** 3, ("тысяча", "тысячи", "тысяч"), True),
    ]
    words = []
    for div, forms, fem in scales:
        q, n = divmod(n, div)
        if q:
            words.append(_triple(q, feminine=fem))
            words.append(_plural(q, *forms))
    if n:
        words.append(_triple(n, feminine=feminine))
    return " ".join(w for w in words if w)


def rub_propisyu(amount):
    """150000 -> 'Сто пятьдесят тысяч рублей 00 копеек'."""
    amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    rub = int(amount)
    kop = int((amount - rub) * 100)
    s = chislo_propisyu(rub)
    s = s[:1].upper() + s[1:]
    return f"{s} {_plural(rub, 'рубль', 'рубля', 'рублей')} {kop:02d} {_plural(kop, 'копейка', 'копейки', 'копеек')}"


def money_fmt(amount):
    """150000 -> '150 000,00'"""
    amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = f"{amount:,.2f}".replace(",", " ").replace(".", ",")
    return s


_UNITS_G = ["", "одного", "двух", "трёх", "четырёх", "пяти", "шести", "семи", "восьми", "девяти"]
_TEENS_G = ["десяти", "одиннадцати", "двенадцати", "тринадцати", "четырнадцати",
            "пятнадцати", "шестнадцати", "семнадцати", "восемнадцати", "девятнадцати"]
_TENS_G  = ["", "", "двадцати", "тридцати", "сорока", "пятидесяти",
            "шестидесяти", "семидесяти", "восьмидесяти", "девяноста"]


def dnej_propisyu(n):
    """Число в родительном падеже для оборотов 'в течение 5 (пяти) дней', 1..100."""
    n = int(n)
    if n == 100:
        return "ста"
    if 10 <= n <= 19:
        return _TEENS_G[n - 10]
    t, u = divmod(n, 10)
    parts = []
    if t:
        parts.append(_TENS_G[t])
    if u:
        parts.append(_UNITS_G[u])
    return " ".join(parts) if parts else "ноля"


_MONTHS_G = ["", "января", "февраля", "марта", "апреля", "мая", "июня",
             "июля", "августа", "сентября", "октября", "ноября", "декабря"]


def date_ru(d):
    """date -> '«11» июня 2026 г.'"""
    if isinstance(d, str):
        return d
    return f"«{d.day:02d}» {_MONTHS_G[d.month]} {d.year} г."


# ----------------------------------------------------------------------------
# 2. РЕНДЕР DOCX-ШАБЛОНОВ
# ----------------------------------------------------------------------------
_RX_TAG = re.compile(r"(\{[\{%])(.*?)([\}%]\})", re.DOTALL)
_RX_RUNBREAK = re.compile(r"</w:t>.*?<w:t(?: [^>]*)?>", re.DOTALL)
_RX_P_TAG = re.compile(
    r"<w:p\b(?:(?!</w:p>).)*?\{%p((?:(?!%\}).)*?)%\}(?:(?!</w:p>).)*?</w:p>", re.DOTALL)
_RX_TR_TAG = re.compile(
    r"<w:tr\b(?:(?!</w:tr>).)*?\{%tr((?:(?!%\}).)*?)%\}(?:(?!</w:tr>).)*?</w:tr>", re.DOTALL)


def _merge_split_tags(xml):
    """Word при редактировании может разрезать {{ тег }} на несколько runs —
    склеиваем содержимое тегов обратно."""
    def fix(m):
        inner = _RX_RUNBREAK.sub("", m.group(2))
        return m.group(1) + inner + m.group(3)
    return _RX_TAG.sub(fix, xml)


def _prepare_xml(xml):
    xml = _merge_split_tags(xml)
    xml = _RX_TR_TAG.sub(lambda m: "{%" + m.group(1) + "%}", xml)   # строки таблиц
    xml = _RX_P_TAG.sub(lambda m: "{%" + m.group(1) + "%}", xml)    # абзацы-маркеры
    return xml


def _finalize(value):
    """Экранируем XML и превращаем \n в перенос строки Word."""
    s = _xml_escape(str(value))
    return s.replace("\n", '</w:t><w:br/><w:t xml:space="preserve">')


_env = jinja2.Environment(undefined=jinja2.StrictUndefined, finalize=_finalize,
                          autoescape=False)


def render_docx(template_path, context, out_path):
    """Заполняет docx-шаблон значениями context и сохраняет результат."""
    with zipfile.ZipFile(template_path) as zin:
        items = {n: zin.read(n) for n in zin.namelist()}
    targets = [n for n in items if re.fullmatch(
        r"word/(document|header\d*|footer\d*)\.xml", n)]
    for name in targets:
        xml = items[name].decode("utf-8")
        xml = _prepare_xml(xml)
        try:
            xml = _env.from_string(xml).render(**context)
        except jinja2.UndefinedError as e:
            raise RuntimeError(f"В шаблоне «{os.path.basename(template_path)}» "
                               f"есть переменная без значения: {e}") from e
        items[name] = xml.encode("utf-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in items.items():
            zout.writestr(name, data)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with open(out_path, "wb") as f:
        f.write(buf.getvalue())
    return out_path


# ----------------------------------------------------------------------------
# 3. КОНТЕКСТ ДОГОВОРА (поля + вычисляемые значения)
# ----------------------------------------------------------------------------
def rekvizity_text(p):
    """Многострочный блок реквизитов стороны для раздела «Адреса и реквизиты»."""
    lines = [p.get("наименование", "")]
    if p.get("инн"):
        s = "ИНН " + p["инн"]
        if p.get("кпп"):
            s += ", КПП " + p["кпп"]
        lines.append(s)
    if p.get("огрн"):
        lines.append(("ОГРНИП " if p.get("тип") == "ИП" else "ОГРН ") + p["огрн"])
    if p.get("паспорт"):
        lines.append("Паспорт: " + p["паспорт"])
    if p.get("адрес"):
        lines.append("Адрес: " + p["адрес"])
    if p.get("адрес_корр"):
        lines.append("Адрес для корреспонденции: " + p["адрес_корр"])
    if p.get("счет"):
        lines.append("Р/с " + p["счет"])
    if p.get("банк"):
        lines.append("Банк: " + p["банк"])
    if p.get("бик"):
        s = "БИК " + p["бик"]
        if p.get("корсчет"):
            s += ", к/с " + p["корсчет"]
        lines.append(s)
    if p.get("телефон"):
        lines.append("Тел.: " + p["телефон"])
    if p.get("email"):
        lines.append("E-mail: " + p["email"])
    return "\n".join(l for l in lines if l)


def short_fio(fio):
    """'Иванов Иван Иванович' -> 'Иванов И.И.'"""
    parts = str(fio).split()
    if len(parts) >= 2:
        return parts[0] + " " + ".".join(w[0] for w in parts[1:]) + "."
    return str(fio)


def build_context(d):
    """d — словарь данных из формы. Возвращает контекст для шаблонов."""
    зак, исп = d["заказчик"], d["исполнитель"]
    исп_тип = исп.get("тип") or "Самозанятый"
    самозанятый = исп_тип != "ИП"

    # услуги: сумма строки = кол-во × цена (если цена не задана — берем «стоимость»)
    услуги_ctx, итого = [], Decimal("0")
    for i, u in enumerate(u for u in d.get("услуги", [])
                          if str(u.get("наименование", "")).strip()):
        колво = Decimal(str(u.get("колво") or 1))
        цена = u.get("цена")
        if цена in (None, ""):
            цена = u.get("стоимость") or 0
        цена = Decimal(str(цена))
        сумма = (цена * колво).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        итого += сумма
        услуги_ctx.append({
            "номер": i + 1,
            "наименование": str(u.get("наименование", "")),
            "колво": f"{колво.normalize():f}".rstrip("0").rstrip(".") or "0",
            "цена": money_fmt(цена),
            "стоимость": money_fmt(сумма),
            "начало": date_ru(u.get("начало") or d["услуги_начало"]),
            "окончание": date_ru(u.get("окончание") or d["услуги_окончание"]),
            "требования": str(u.get("требования") or "—"),
        })
    стоимость = Decimal(str(d.get("стоимость") or 0)) or итого

    # сквозная нумерация разделов с учётом включённых блоков
    n = 6  # последний фиксированный раздел — «6. Ответственность Сторон»
    sec = {}
    if d.get("блок_нда"):
        n += 1; sec["нда"] = n
    if d.get("блок_фм"):
        n += 1; sec["фм"] = n
    sec["раст"], sec["спор"], sec["проч"], sec["прил"], sec["рекв"] = \
        n + 1, n + 2, n + 3, n + 4, n + 5

    дни = {k: int(d.get(k, v)) for k, v in {
        "справка_дней": 3, "пояснения_дней": 3, "акт_дней": 5, "приемка_дней": 5,
        "устранение_дней": 5, "предоплата_дней": 5, "постоплата_дней": 5,
        "расторжение_дней": 10, "претензия_дней": 10, "фм_уведомление_дней": 10}.items()}

    ctx = {
        "договор_номер": d["договор_номер"],
        "договор_дата": date_ru(d["договор_дата"]),
        "город": d.get("город", ""),
        "услуги_начало": date_ru(d["услуги_начало"]),
        "услуги_окончание": date_ru(d["услуги_окончание"]),

        "заказчик_наименование": зак.get("наименование", ""),
        "заказчик_должность_род": зак.get("должность_род", "Генерального директора"),
        "заказчик_подписант_род": зак.get("подписант_род", ""),
        "заказчик_основание_род": зак.get("основание_род", "Устава"),
        "заказчик_подписант_кратко": short_fio(зак.get("подписант", зак.get("подписант_род", ""))),
        "заказчик_инн": зак.get("инн", ""), "заказчик_кпп": зак.get("кпп", ""),
        "заказчик_огрн": зак.get("огрн", ""),
        "заказчик_адрес": зак.get("адрес", ""),
        "заказчик_реквизиты": rekvizity_text(зак),

        "исполнитель_фио": исп.get("наименование", ""),
        "исполнитель_фио_кратко": short_fio(исп.get("наименование", "")),
        "исполнитель_тип": исп_тип,
        "исполнитель_инн": исп.get("инн", ""),
        "исполнитель_огрн": исп.get("огрн", ""),
        "исполнитель_паспорт": исп.get("паспорт", ""),
        "исполнитель_адрес": исп.get("адрес", ""),
        "исполнитель_банк": исп.get("банк", ""), "исполнитель_бик": исп.get("бик", ""),
        "исполнитель_счет": исп.get("счет", ""), "исполнитель_корсчет": исп.get("корсчет", ""),
        "исполнитель_реквизиты": rekvizity_text(исп),

        "стоимость_цифрами": money_fmt(стоимость),
        "стоимость_прописью": rub_propisyu(стоимость),
        "предоплата_процент": d.get("предоплата_процент", 50),
        "постоплата_процент": d.get("постоплата_процент", 50),
        "штраф_чек_процент": d.get("штраф_чек_процент", 20),
        "штраф_уведомление": money_fmt(d.get("штраф_уведомление", 10000)),
        "штраф_уведомление_проп": rub_propisyu(d.get("штраф_уведомление", 10000)),

        "блок_нда": bool(d.get("блок_нда")),
        "блок_фм": bool(d.get("блок_фм")),
        "блок_согласие": bool(d.get("блок_согласие")),
        "блок_ис": bool(d.get("блок_ис", True)),
        "нда_лет": int(d.get("нда_лет", 3)),
        "нда_лет_проп": chislo_propisyu(int(d.get("нда_лет", 3))),
        "фм_месяцев": int(d.get("фм_месяцев", 2)),
        "фм_месяцев_проп": chislo_propisyu(int(d.get("фм_месяцев", 2))),

        "услуги": услуги_ctx,

        "акт_номер": d.get("акт_номер") or d["договор_номер"],
        "акт_дата": date_ru(d.get("акт_дата") or d["услуги_окончание"]),
        "счет_номер": d.get("счет_номер") or d["договор_номер"],
        "счет_дата": date_ru(d.get("счет_дата") or d["договор_дата"]),
        "согласие_до": date_ru(d.get("согласие_до") or d["услуги_окончание"]),

        # --- поля документов «Точка» -------------------------------------
        "прил_номер": d.get("прил_номер") or "1",
        "прил_дата": date_ru(d.get("прил_дата") or d["договор_дата"]),
        "оферта_срок_оплаты": d.get("оферта_срок_оплаты")
            or "в течение 5 (пяти) рабочих дней с даты выставления Счета",
        "оферта_срок_работ": d.get("оферта_срок_работ")
            or "в течение 10 (десяти) рабочих дней с даты внесения аванса",
        "оферта_результат": d.get("оферта_результат")
            or "результат работ, указанных в Счете",
        "оферта_формат": d.get("оферта_формат") or "ссылкой на облачное хранилище",
        "ндс_строка": d.get("ндс_строка")
            or ("без НДС (Подрядчик применяет НПД)" if самозанятый else "без НДС"),
        "ндс_предложение": d.get("ндс_предложение")
            or ("НДС не облагается, поскольку Подрядчик является плательщиком налога "
                "на профессиональный доход." if самозанятый else "НДС не облагается."),
    }
    for k, v in дни.items():
        ctx[k] = v
        ctx[k + "_проп"] = dnej_propisyu(v)
    for k, v in sec.items():
        ctx["с_" + k] = v
    return ctx


# ----------------------------------------------------------------------------
# 4. ГЕНЕРАЦИЯ ПАКЕТА ДОКУМЕНТОВ
# ----------------------------------------------------------------------------
TPL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")

DOCS = {  # ключ: (файл шаблона, имя результата)
    "договор":  ("dogovor.docx",      "Договор_№{n}.docx"),
    "прил1":    ("pril1_perechen.docx", "Приложение_1_Перечень_услуг.docx"),
    "прил2":    ("pril2_zadanie.docx",  "Приложение_2_Задание.docx"),
    "акт":      ("akt.docx",          "Приложение_3_Акт.docx"),
    "согласие": ("soglasie_pd.docx",  "Приложение_4_Согласие_ПД.docx"),
    "счет":     ("schet.docx",        "Счет_№{n}.docx"),
    # документы по формам «Точка»
    "точка_прил": ("tochka_prilozhenie.docx", "Точка_Приложение_к_договору.docx"),
    "точка_акт":  ("tochka_akt.docx",         "Точка_Акт_сдачи-приемки.docx"),
    "точка_счет": ("tochka_schet_oferta.docx", "Точка_Счет-оферта_№{n}.docx"),
}


def generate_package(d, out_dir, docs=None, tpl_dir=None):
    """Генерирует выбранные документы, возвращает список путей.
    tpl_dir — папка с шаблонами (по умолчанию локальная templates/)."""
    ctx = build_context(d)
    docs = docs or ["договор", "прил1", "прил2", "акт"]
    tpl_dir = tpl_dir or TPL_DIR
    if "согласие" in docs:
        ctx["блок_согласие"] = True
    paths = []
    for key in docs:
        tpl, name = DOCS[key]
        out = os.path.join(out_dir, name.format(n=str(d["договор_номер"]).replace("/", "-")))
        paths.append(render_docx(os.path.join(tpl_dir, tpl), ctx, out))
    return paths


# ----------------------------------------------------------------------------
# 5. БАЗА КОНТРАГЕНТОВ (Excel)
# ----------------------------------------------------------------------------
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "контрагенты.xlsx")
PARTY_COLS = ["тип", "наименование", "инн", "кпп", "огрн", "адрес",
              "должность_род", "подписант_род", "подписант", "основание_род",
              "паспорт", "счет", "банк", "бик", "корсчет", "телефон", "email"]
JOURNAL_COLS = ["создано", "договор_номер", "договор_дата", "заказчик",
                "исполнитель", "сумма", "блоки", "документы"]


def _open_db():
    import openpyxl
    if os.path.exists(DB_PATH):
        return openpyxl.load_workbook(DB_PATH)
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Контрагенты"; ws.append(PARTY_COLS)
    wj = wb.create_sheet("Журнал"); wj.append(JOURNAL_COLS)
    return wb


def load_parties():
    if not os.path.exists(DB_PATH):
        return []
    wb = _open_db(); ws = wb["Контрагенты"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return [dict(zip(PARTY_COLS, [("" if v is None else str(v)) for v in r]))
            for r in rows if any(r)]


def save_party(p):
    """Добавляет контрагента или обновляет существующего (по ИНН/наименованию)."""
    wb = _open_db(); ws = wb["Контрагенты"]
    key = p.get("инн") or p.get("наименование")
    if not key:
        return
    for row in ws.iter_rows(min_row=2):
        rk = (row[2].value or row[1].value or "")
        if str(rk) == str(key):
            for i, c in enumerate(PARTY_COLS):
                row[i].value = p.get(c, "")
            wb.save(DB_PATH); return
    ws.append([p.get(c, "") for c in PARTY_COLS])
    wb.save(DB_PATH)


def append_journal(d, docs):
    wb = _open_db(); ws = wb["Журнал"]
    сумма = d.get("стоимость") or sum(
        float(u.get("цена") or u.get("стоимость") or 0) * float(u.get("колво") or 1)
        for u in d.get("услуги", []))
    блоки = ", ".join(b for b, on in [("НДА", d.get("блок_нда")),
                                      ("Форс-мажор", d.get("блок_фм")),
                                      ("Согласие ПД", "согласие" in docs)] if on) or "—"
    ws.append([datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
               str(d["договор_номер"]), date_ru(d["договор_дата"]),
               d["заказчик"].get("наименование", ""), d["исполнитель"].get("наименование", ""),
               money_fmt(сумма), блоки, ", ".join(docs)])
    wb.save(DB_PATH)


def next_contract_number():
    if not os.path.exists(DB_PATH):
        return "1"
    wb = _open_db(); ws = wb["Журнал"]
    nums = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        m = re.match(r"\d+", str(r[1] or ""))
        if m:
            nums.append(int(m.group()))
    return str(max(nums) + 1) if nums else "1"
